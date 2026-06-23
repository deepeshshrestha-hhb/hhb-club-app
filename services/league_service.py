import openpyxl
from collections import Counter, defaultdict
from pathlib import Path
from config import Config
from services.tournament_service import _clean, _fmt_date

TOURNAMENTS_DIR = Path(Config.BASE_DIR) / "tournaments"


def list_league_years():
    years = []
    if TOURNAMENTS_DIR.exists():
        for f in TOURNAMENTS_DIR.glob("HHB Annual Players League - *.xlsm"):
            try:
                years.append(int(f.stem.split("-")[-1].strip()))
            except ValueError:
                pass
    return sorted(years, reverse=True)


def get_league(year):
    path = TOURNAMENTS_DIR / f"HHB Annual Players League - {year}.xlsm"
    if not path.exists():
        return None

    wb = openpyxl.load_workbook(path, data_only=True)

    # Rules
    rules = []
    for row in range(2, 30):
        text = wb["PointRules"].cell(row, 2).value
        if text:
            rules.append(_clean(text))

    match_ws = wb[str(year)]

    # --- Matches ---
    matches = []
    min_date = None
    max_date = None

    for row in range(5, 700):
        date_val = match_ws.cell(row, 1).value
        no = match_ws.cell(row, 2).value
        if no is None or not hasattr(date_val, "year"):
            continue

        p1 = _clean(match_ws.cell(row, 3).value)
        p2 = _clean(match_ws.cell(row, 4).value)
        s1 = match_ws.cell(row, 5).value  # team 1 score
        p3 = _clean(match_ws.cell(row, 6).value)
        p4 = _clean(match_ws.cell(row, 7).value)
        s2 = match_ws.cell(row, 8).value  # team 2 score
        w1 = _clean(match_ws.cell(row, 9).value)
        w2 = _clean(match_ws.cell(row, 10).value)
        diff = match_ws.cell(row, 11).value

        if s1 is None or s2 is None:
            continue

        if min_date is None or date_val < min_date:
            min_date = date_val
        if max_date is None or date_val > max_date:
            max_date = date_val

        is_deuce = int(s1) == 21 and int(s2) == 20 or int(s1) == 20 and int(s2) == 21

        matches.append({
            "no": no,
            "date": _fmt_date(date_val),
            "date_raw": date_val,
            "p1": p1, "p2": p2,
            "score1": int(s1),
            "p3": p3, "p4": p4,
            "score2": int(s2),
            "winner": f"{w1} & {w2}" if w1 and w2 else w1 or w2,
            "diff": abs(int(diff)) if diff is not None else abs(int(s1) - int(s2)),
            "is_deuce": is_deuce,
            "players": f"{p1}|{p2}|{p3}|{p4}",
        })

    # --- Standings ---
    standings = []
    for row in range(3, 100):
        rank = match_ws.cell(row, 19).value
        player = match_ws.cell(row, 20).value
        if rank is None or player is None or not isinstance(rank, (int, float)):
            break
        p = _clean(player)
        pl = match_ws.cell(row, 21).value or 0
        w = match_ws.cell(row, 22).value or 0
        lo = match_ws.cell(row, 23).value or 0
        pts = match_ws.cell(row, 24).value or 0
        standings.append({
            "rank": int(rank),
            "player": p,
            "wins": int(w),
            "played": int(pl),
            "losses": int(lo),
            "points": int(pts),
            "win_pct": round(w / pl * 100) if pl else 0,
        })

    # --- Analytics ---
    total = len(matches)
    sundays = len({m["date_raw"] for m in matches})
    deuce = [m for m in matches if m["is_deuce"]]
    diffs = [m["diff"] for m in matches if m["diff"] > 0]
    avg_diff = round(sum(diffs) / len(diffs), 1) if diffs else 0
    biggest = max(matches, key=lambda m: m["diff"]) if matches else None
    squeaky = [m for m in matches if m["diff"] == 1]

    # Score frequency
    score_counter = Counter()
    for m in matches:
        hi, lo = max(m["score1"], m["score2"]), min(m["score1"], m["score2"])
        score_counter[(hi, lo)] += 1
    top_scores = [(f"{h}-{l}", c) for (h, l), c in score_counter.most_common(5)]

    # Max wins by individual on a single day — Top 3
    day_wins = defaultdict(int)  # (player, date) -> wins
    for m in matches:
        for w in [m["p1"] if m["winner"] == f"{m['p1']} & {m['p2']}" else None,
                  m["p2"] if m["winner"] == f"{m['p1']} & {m['p2']}" else None,
                  m["p3"] if m["winner"] == f"{m['p3']} & {m['p4']}" else None,
                  m["p4"] if m["winner"] == f"{m['p3']} & {m['p4']}" else None]:
            if w:
                day_wins[(w, m["date"])] += 1
    top_day_wins = sorted(day_wins.items(), key=lambda x: -x[1])[:3]
    top_individual_day = [{"player": p, "date": d, "wins": w} for (p, d), w in top_day_wins]

    # Pair wins/losses across full season
    pair_wins = defaultdict(int)
    pair_losses = defaultdict(int)
    for m in matches:
        pair1 = tuple(sorted([m["p1"], m["p2"]]))
        pair2 = tuple(sorted([m["p3"], m["p4"]]))
        winner_pair = tuple(sorted([w.strip() for w in m["winner"].split("&")])) if "&" in m["winner"] else None
        if winner_pair == pair1:
            pair_wins[pair1] += 1
            pair_losses[pair2] += 1
        elif winner_pair == pair2:
            pair_wins[pair2] += 1
            pair_losses[pair1] += 1

    all_pairs = set(list(pair_wins.keys()) + list(pair_losses.keys()))
    pair_records = []
    for pair in all_pairs:
        w = pair_wins.get(pair, 0)
        l = pair_losses.get(pair, 0)
        if w + l < 2:
            continue
        pair_records.append({
            "pair": f"{pair[0]} & {pair[1]}",
            "wins": w,
            "losses": l,
            "played": w + l,
            "undefeated": l == 0 and w >= 5,
            "win_pct": round(w / (w + l) * 100) if (w + l) else 0,
        })
    pair_records.sort(key=lambda x: (-x["wins"], x["losses"]))
    top_pairs = pair_records[:10]
    undefeated_pairs = [p for p in pair_records if p["undefeated"]]

    # Matches per Sunday
    by_date = defaultdict(int)
    for m in matches:
        by_date[m["date_raw"]] += 1
    busiest = max(by_date.items(), key=lambda x: x[1]) if by_date else None

    # Player appearances (for filter dropdown)
    player_set = set()
    for m in matches:
        for p in [m["p1"], m["p2"], m["p3"], m["p4"]]:
            if p:
                player_set.add(p)
    all_players = sorted(player_set)

    analytics = {
        "total_matches": total,
        "total_sundays": sundays,
        "avg_per_sunday": round(total / sundays, 1) if sundays else 0,
        "deuce_count": len(deuce),
        "deuce_pct": round(len(deuce) / total * 100, 1) if total else 0,
        "avg_diff": avg_diff,
        "biggest_win": biggest,
        "squeaky_wins": len(squeaky),
        "top_scores": top_scores,
        "busiest_sunday": {"date": _fmt_date(busiest[0]), "matches": busiest[1]} if busiest else None,
        "top_individual_day": top_individual_day,
        "top_pairs": top_pairs,
        "undefeated_pairs": undefeated_pairs,
    }

    return {
        "year": year,
        "title": f"HHB Annual Players League {year}",
        "season_start": _fmt_date(min_date),
        "season_end": _fmt_date(max_date),
        "standings": standings,
        "winner": standings[0]["player"] if standings else "",
        "runner_up": standings[1]["player"] if len(standings) > 1 else "",
        "third": standings[2]["player"] if len(standings) > 2 else "",
        "matches": matches,
        "all_players": all_players,
        "analytics": analytics,
        "rules": rules,
    }
