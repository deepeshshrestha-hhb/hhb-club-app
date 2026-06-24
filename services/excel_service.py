import io
import zipfile

import openpyxl
import pandas as pd
from pathlib import Path
from config import Config


DATA_DIR = Path(Config.DATA_DIR)


def _rebuild_with_forward_slashes(path, **kwargs):
    """Rebuild an OOXML package in memory, replacing backslash zip-entry
    separators with forward slashes, then load it with openpyxl.

    Some of the club's .xlsm scoresheets were saved by a tool that stored entry
    names like ``xl\\sharedStrings.xml``. The OOXML/ZIP spec requires forward
    slashes, and a package's own relationships reference ``xl/sharedStrings.xml``.
    Windows' zipfile silently maps ``\\`` -> ``/`` so these load fine locally, but
    on Linux the literal backslash names don't match and openpyxl raises
    ``KeyError: 'xl/sharedStrings.xml'``. Forward-slash names fix that.
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(path) as zin, \
            zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            info = zipfile.ZipInfo(
                item.filename.replace("\\", "/"), date_time=item.date_time
            )
            info.compress_type = item.compress_type
            info.external_attr = item.external_attr
            info.internal_attr = item.internal_attr
            info.create_system = item.create_system
            zout.writestr(info, data)
    buf.seek(0)
    return openpyxl.load_workbook(buf, **kwargs)


def load_workbook_normalized(path, **kwargs):
    """Like ``openpyxl.load_workbook`` but tolerant of .xlsm/.xlsx packages whose
    zip entries use backslash path separators (which otherwise fail on Linux with
    ``KeyError: 'xl/sharedStrings.xml'``). Normal/correct files take the fast path
    unchanged; only files that raise are rebuilt with forward slashes."""
    try:
        return openpyxl.load_workbook(path, **kwargs)
    except KeyError:
        return _rebuild_with_forward_slashes(path, **kwargs)


def ensure_data_dir():
    DATA_DIR.mkdir(exist_ok=True)


def load_excel(filename, **kwargs):
    ensure_data_dir()
    path = DATA_DIR / filename
    if not path.exists():
        return pd.DataFrame()
    return pd.read_excel(path, **kwargs)


def save_excel(df, filename):
    ensure_data_dir()
    path = DATA_DIR / filename
    df.to_excel(path, index=False)
    # Push the durable copy back to R2 (no-op when R2 isn't configured).
    from services import r2_service
    r2_service.upload_file(path)
