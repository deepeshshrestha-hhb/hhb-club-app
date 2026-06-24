"""
Two things live in this file:

1. The original generic tournament CRUD (backed by Tournaments.xlsx) used by
   the "Create Tournament" form on /tournaments.

2. A read-only parser for the HHB Annual Doubles Classic scoresheets
   (.xlsm files in the 'tournaments' folder), used by the /tournaments/doubles
   archive pages.

NOTE on the Doubles parser: it's built against the modern tournament template
used in recent years (2 groups of any size, top 4 from each group go to
Quarters -> Semis -> Final/3rd place). Older formats (e.g. 2018/2019 single
round-robin league, or 2022's 3-group "Super 6/Super 3" stage) use a different
layout and are not yet supported - they'll need their own handling later.
"""
import openpyxl
import pandas as pd
from pathlib import Path
from config import Config
from services.excel_service import load_excel, save_excel


# --- Generic tournament CRUD (Tournaments.xlsx) ---

def _load_tournaments_df():
    df = load_excel("Tournaments.xlsx")
    if df.empty:
        df = pd.DataFrame(columns=["Id", "Name", "Date"])
    return df


def get_all_tournaments():
    df = _load_tournaments_df()
    tournaments = []
    for _, row in df.iterrows():
        tournaments.append(
            {
                "id": int(row["Id"]),
                "name": row["Name"],
                "date": str(row["Date"]),
            }
        )
    return tournaments


def get_tournament_by_id(tournament_id: int):
    df = _load_tournaments_df()
    row = df[df["Id"] == tournament_id]
    if row.empty:
        return None
    r = row.iloc[0]
    return {
        "id": int(r["Id"]),
        "name": r["Name"],
        "date": str(r["Date"]),
    }


def create_tournament(name: str, date: str):
    df = _load_tournaments_df()
    new_id = 1 if df.empty else int(df["Id"].max()) + 1
    new_row = {"Id": new_id, "Name": name, "Date": date}
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    save_excel(df, "Tournaments.xlsx")


# --- HHB Annual Doubles Classic archive (read-only Excel parser) ---

TOURNAMENTS_DIR = Path(Config.BASE_DIR) / "tournaments"

# Years confirmed to use the modern 2-group / QF-SF-Final template.
# Extend this as more years are checked and confirmed to fit the same layout.
SUPPORTED_DOUBLES_YEARS = [2018, 2021, 2023, 2024, 2025, 2026]


def _doubles_filename(year):
    return f"HHB Annual Doubles Classic - {year}.xlsm"


def list_doubles_tournament_years():
    """Years we have a scoresheet for, regardless of whether the parser supports them yet."""
    years = []
    if TOURNAMENTS_DIR.exists():
        for f in TOURNAMENTS_DIR.glob("HHB Annual Doubles Classic - *.xlsm"):
            try:
                year = int(f.stem.split("-")[-1].strip())
                years.append(year)
            except ValueError:
                continue
    return sorted(years, reverse=True)


def _clean(value):
    if value is None:
        return ""
    text = str(value)
    return "" if text.lower() == "nan" else text


def _fmt_date(value):
    if hasattr(value, "strftime"):
        return value.strftime("%d %b %Y")
    return _clean(value)


def _parse_rules(ws):
    rules = []
    row = 3
    while True:
        num = ws[f"G{row}"].value
        text = ws[f"H{row}"].value
        if num is None and text is None:
            break
        if text:
            rules.append(_clean(text))
        row += 1
    return rules


def _parse_important_dates(ws):
    dates = []
    row = 4
    while True:
        label = ws[f"B{row}"].value
        date_val = ws[f"E{row}"].value
        if label is None:
            break
        dates.append({"label": _clean(label), "date": _fmt_date(date_val)})
        row += 1
    return dates


def _find_standings_start_row(ws, max_search=6):
    """The points/standings table header row shifts between tournament years
    (row 2 in some years, row 1 in others). Instead of assuming a fixed row,
    find the row where rank '1' actually appears in column J."""
    for row in range(1, max_search + 1):
        if ws[f"J{row}"].value == 1:
            return row
    return 3  # sensible fallback matching the modern template


def _parse_group_sheet(ws, group_name):
    # Matches table: columns A-H starting row 3
    matches = []
    row = 3
    while ws[f"A{row}"].value is not None:
        matches.append(
            {
                "no": ws[f"A{row}"].value,
                "date": _fmt_date(ws[f"B{row}"].value),
                "team1": _clean(ws[f"C{row}"].value),
                "score1": ws[f"D{row}"].value,
                "team2": _clean(ws[f"E{row}"].value),
                "score2": ws[f"F{row}"].value,
                "winner": _clean(ws[f"G{row}"].value),
                "margin": ws[f"H{row}"].value,
            }
        )
        row += 1

    # Points table: columns J-R, starting row varies by year - detect it.
    standings = []
    row = _find_standings_start_row(ws)
    while ws[f"K{row}"].value is not None:
        standings.append(
            {
                "rank": ws[f"J{row}"].value,
                "team": _clean(ws[f"K{row}"].value),
                "played": ws[f"L{row}"].value,
                "won": ws[f"M{row}"].value,
                "lost": ws[f"N{row}"].value,
                "points": ws[f"O{row}"].value,
                "pf": ws[f"P{row}"].value,
                "pa": ws[f"Q{row}"].value,
                "nsd": ws[f"R{row}"].value,
            }
        )
        row += 1
    # Already in points order in the sheet, but sort defensively just in case.
    standings.sort(key=lambda t: (-(t["points"] or 0), -(t["nsd"] or 0)))

    return {"name": group_name, "matches": matches, "standings": standings}


def _parse_knockouts_direct_semis(ws):
    """
    4-team direct-to-semis format used in 2019/2021 (no quarterfinal round -
    group winners/runners-up go straight to semis). Some matches were decided
    by forfeit/walkover rather than scores, so where available we trust the
    sheet's own recorded Finalist/Winner/Runner Up/Third cells (which already
    account for that) rather than only recomputing from raw set scores.
    """
    def forfeit_note(date_cell_value):
        # The "date" cell below each match sometimes holds text like
        # "FORFEIT" instead of an actual date.
        if isinstance(date_cell_value, str) and date_cell_value.strip():
            return date_cell_value.strip()
        return ""

    sf1_team1, sf1_team1_sets = _clean(ws["B3"].value), _sets(ws, "D3", "F3", "H3")
    sf1_team2, sf1_team2_sets = _clean(ws["B9"].value), _sets(ws, "D9", "F9", "H9")
    sf2_team1, sf2_team1_sets = _clean(ws["B15"].value), _sets(ws, "D15", "F15", "H15")
    sf2_team2, sf2_team2_sets = _clean(ws["B21"].value), _sets(ws, "D21", "F21", "H21")

    sf1_note = forfeit_note(ws["B7"].value)
    sf2_note = forfeit_note(ws["B19"].value)

    # Prefer the sheet's own recorded Finalist name (J5/J17) - it already
    # reflects forfeits/walkovers correctly. Fall back to computing from sets.
    sf1_winner = _clean(ws["J5"].value) or _match_winner(sf1_team1, sf1_team1_sets, sf1_team2, sf1_team2_sets)
    sf2_winner = _clean(ws["J17"].value) or _match_winner(sf2_team1, sf2_team1_sets, sf2_team2, sf2_team2_sets)
    sf1_loser = sf1_team2 if sf1_winner == sf1_team1 else sf1_team1 if sf1_winner else ""
    sf2_loser = sf2_team2 if sf2_winner == sf2_team1 else sf2_team1 if sf2_winner else ""

    sf1 = {
        "team1": sf1_team1, "team1_sets": sf1_team1_sets,
        "team2": sf1_team2, "team2_sets": sf1_team2_sets,
        "winner": sf1_winner, "note": sf1_note,
    }
    sf2 = {
        "team1": sf2_team1, "team1_sets": sf2_team1_sets,
        "team2": sf2_team2, "team2_sets": sf2_team2_sets,
        "winner": sf2_winner, "note": sf2_note,
    }

    final_team1_sets = _sets(ws, "L5", "N5", "P5")
    final_team2_sets = _sets(ws, "L17", "N17", "P17")
    # Prefer the sheet's own recorded Winner cell (R11) over recomputing.
    final_winner = _clean(ws["R11"].value) or _match_winner(
        sf1_winner, final_team1_sets, sf2_winner, final_team2_sets
    )
    final = {
        "team1": sf1_winner, "team1_sets": final_team1_sets,
        "team2": sf2_winner, "team2_sets": final_team2_sets,
        "winner": final_winner,
    }
    runner_up = _clean(ws["R21"].value) or (
        final["team2"] if final_winner == final["team1"] else final["team1"] if final_winner else ""
    )

    third_team1_sets = _sets(ws, "L27", "N27", "P27")
    third_team2_sets = _sets(ws, "L34", "N34", "P34")
    third_winner = _clean(ws["R30"].value) or _match_winner(
        sf1_loser, third_team1_sets, sf2_loser, third_team2_sets
    )
    third_place = {
        "team1": sf1_loser, "team1_sets": third_team1_sets,
        "team2": sf2_loser, "team2_sets": third_team2_sets,
        "winner": third_winner,
    }

    return {
        "quarterfinals": [],
        "semifinals": [sf1, sf2],
        "final": final,
        "third_place": third_place,
        "winner": final_winner,
        "runner_up": runner_up,
        "third": third_winner,
    }


def _parse_knockouts_2022(ws):
    """
    One-off bespoke layout for 2022: qualification came via 'Super 6'/'Super 3'
    placement rather than group position, and one slot was decided by a
    walkover ('Challenger Match'). The bracket path here is reconstructed on a
    best-effort basis, but the Winner/Runner-up/Third labels are stored as
    plain text in the sheet (not formulas), so those three results are reliable
    even if the path to get there is approximate.
    """
    sf1 = {
        "team1": _clean(ws["G3"].value), "team1_sets": _sets(ws, "I3", "K3", "M3"),
        "team2": _clean(ws["G9"].value), "team2_sets": _sets(ws, "I9", "K9", "M9"),
    }
    sf1["winner"] = _match_winner(sf1["team1"], sf1["team1_sets"], sf1["team2"], sf1["team2_sets"])

    # SF2: Usman/Prasanna (G15) vs the Challenger Match team (B18, won by walkover)
    challenger_team = _clean(ws["B18"].value)
    challenger_note = _clean(ws["D18"].value)  # e.g. "WIN BY DEFAULT"
    sf2 = {
        "team1": _clean(ws["G15"].value), "team1_sets": _sets(ws, "I15", "K15", "M15"),
        "team2": challenger_team, "team2_sets": [],
        "note": challenger_note,
    }
    # Walkover - winner is whichever team isn't credited with the walkover loss.
    # The finalist name at O17 tells us who actually progressed.
    sf2["winner"] = _clean(ws["O17"].value) or sf2["team1"]

    final = {
        "team1": _clean(ws["O5"].value), "team1_sets": _sets(ws, "Q5", "S5", "U5"),
        "team2": _clean(ws["O17"].value), "team2_sets": _sets(ws, "Q17", "S17", "U17"),
    }
    final["winner"] = _match_winner(final["team1"], final["team1_sets"], final["team2"], final["team2_sets"])

    third_place = {
        "team1": _clean(ws["O27"].value), "team1_sets": _sets(ws, "Q27", "S27", "U27"),
        "team2": _clean(ws["O34"].value), "team2_sets": _sets(ws, "Q34", "S34", "U34"),
    }
    third_place["winner"] = _match_winner(
        third_place["team1"], third_place["team1_sets"], third_place["team2"], third_place["team2_sets"]
    )

    # These are stored as plain text in the sheet (not formulas) - trust them
    # directly over the reconstructed bracket above where they might disagree.
    winner = _clean(ws["W11"].value)
    runner_up = _clean(ws["W21"].value)
    third = _clean(ws["W30"].value)

    return {
        "quarterfinals": [],
        "semifinals": [sf1, sf2],
        "final": final,
        "third_place": third_place,
        "winner": winner or final["winner"],
        "runner_up": runner_up,
        "third": third or third_place["winner"],
    }


def _sets(ws, *coords):
    return [ws[c].value for c in coords if ws[c].value is not None]


def _match_winner(team1, team1_sets, team2, team2_sets):
    """Winner is whoever won more sets (not total points)."""
    sets1 = sum(1 for a, b in zip(team1_sets, team2_sets) if a is not None and b is not None and a > b)
    sets2 = sum(1 for a, b in zip(team1_sets, team2_sets) if a is not None and b is not None and b > a)
    if sets1 > sets2:
        return team1
    if sets2 > sets1:
        return team2
    return ""


def _parse_knockouts(ws):
    """
    Fixed-layout parser matching the modern 8-team knockout bracket:
    QF1 (A1 v B4) and QF2 (B2 v A3) feed SF1; QF3 (B1 v A4) and QF4 (A2 v B3)
    feed SF2; SF1 v SF2 feed the Final; SF losers play the 3rd place match.
    """
    def qf(team1_row, team2_row):
        team1 = _clean(ws[f"B{team1_row}"].value)
        team1_sets = _sets(ws, f"D{team1_row}", f"E{team1_row}", f"F{team1_row}")
        team2 = _clean(ws[f"B{team2_row}"].value)
        team2_sets = _sets(ws, f"D{team2_row}", f"E{team2_row}", f"F{team2_row}")
        return {
            "team1": team1,
            "team1_sets": team1_sets,
            "team2": team2,
            "team2_sets": team2_sets,
            "winner": _match_winner(team1, team1_sets, team2, team2_sets),
        }

    def match(team1, team1_sets, team2, team2_sets):
        return {
            "team1": team1,
            "team1_sets": team1_sets,
            "team2": team2,
            "team2_sets": team2_sets,
            "winner": _match_winner(team1, team1_sets, team2, team2_sets),
        }

    qf1 = qf(5, 7)
    qf2 = qf(11, 13)
    qf3 = qf(17, 19)
    qf4 = qf(23, 25)

    sf1 = match(
        _clean(ws["H5"].value), _sets(ws, "J5", "L5", "N5"),
        _clean(ws["H11"].value), _sets(ws, "J11", "L11", "N11"),
    )
    sf2 = match(
        _clean(ws["H17"].value), _sets(ws, "J17", "L17", "N17"),
        _clean(ws["H23"].value), _sets(ws, "J23", "L23", "N23"),
    )
    final = match(
        _clean(ws["P7"].value), _sets(ws, "R7", "T7", "V7"),
        _clean(ws["P19"].value), _sets(ws, "R19", "T19", "V19"),
    )
    third_place = match(
        _clean(ws["P29"].value), _sets(ws, "R29", "T29"),
        _clean(ws["P36"].value), _sets(ws, "R36", "T36"),
    )

    winner = _clean(ws["X13"].value)
    runner_up = _clean(ws["X23"].value)
    third = _clean(ws["X32"].value)

    return {
        "quarterfinals": [qf1, qf2, qf3, qf4],
        "semifinals": [sf1, sf2],
        "final": final,
        "third_place": third_place,
        "winner": winner,
        "runner_up": runner_up,
        "third": third,
    }


def _parse_doubles_2018(wb):
    """
    2018 used a single round-robin league (no groups), an Eliminator match
    to determine the 4th semi-finalist, then best-of-3 knockouts to 21 pts.
    Sheet layout: Dates / League Matches / Knockouts — all different from
    the modern template.
    """
    dates_ws = wb["Dates"]
    league_ws = wb["League Matches"]
    ko_ws = wb["Knockouts"]

    # Important dates: col D = label, col E = date, rows 2-6
    important_dates = []
    for row in range(2, 8):
        label = dates_ws.cell(row, 4).value
        date_val = dates_ws.cell(row, 5).value
        if label:
            important_dates.append({"label": _clean(label), "date": _fmt_date(date_val)})

    # Rules: collect all non-empty col F notes
    rules = []
    for row in range(1, 25):
        note = dates_ws.cell(row, 6).value
        if note:
            rules.append(_clean(note))

    # League matches: data rows start at 3; skip unplayed (no score)
    matches = []
    for row in range(3, 40):
        no = league_ws.cell(row, 2).value
        if no is None:
            break
        score1 = league_ws.cell(row, 4).value
        score2 = league_ws.cell(row, 6).value
        if score1 is None or score2 is None:
            continue
        matches.append({
            "no": no,
            "date": _fmt_date(league_ws.cell(row, 1).value),
            "team1": _clean(league_ws.cell(row, 3).value),
            "score1": score1,
            "team2": _clean(league_ws.cell(row, 5).value),
            "score2": score2,
            "winner": _clean(league_ws.cell(row, 7).value),
            "margin": abs(int(score1) - int(score2)),
        })

    # Standings: rank at col J(10), team at K(11), row 2 = rank 1
    standings = []
    for row in range(2, 12):
        rank = league_ws.cell(row, 10).value
        team = league_ws.cell(row, 11).value
        if rank is None or team is None:
            break
        standings.append({
            "rank": rank,
            "team": _clean(team),
            "played": league_ws.cell(row, 12).value,
            "won": league_ws.cell(row, 13).value,
            "lost": league_ws.cell(row, 14).value,
            "points": league_ws.cell(row, 15).value,
            "pf": league_ws.cell(row, 16).value,
            "pa": league_ws.cell(row, 17).value,
            "nsd": league_ws.cell(row, 18).value,
        })

    # --- Knockouts ---
    # Eliminator: B3/B5 are the two teams; D/F cols are game scores per team
    elim_t1 = _clean(ko_ws.cell(3, 2).value)
    elim_t1_sets = [v for v in [ko_ws.cell(3, 4).value, ko_ws.cell(3, 6).value] if v is not None]
    elim_t2 = _clean(ko_ws.cell(5, 2).value)
    elim_t2_sets = [v for v in [ko_ws.cell(5, 4).value, ko_ws.cell(5, 6).value] if v is not None]

    # SF1: J3 (team1) vs J9 (team2); scores at L/N
    sf1_t1 = _clean(ko_ws.cell(3, 10).value)
    sf1_t1_sets = [v for v in [ko_ws.cell(3, 12).value, ko_ws.cell(3, 14).value] if v is not None]
    sf1_t2 = _clean(ko_ws.cell(9, 10).value)
    sf1_t2_sets = [v for v in [ko_ws.cell(9, 12).value, ko_ws.cell(9, 14).value] if v is not None]
    sf1_winner = _match_winner(sf1_t1, sf1_t1_sets, sf1_t2, sf1_t2_sets)

    # SF2: J15 (team1) vs J21 (team2); scores at L/N/P
    sf2_t1 = _clean(ko_ws.cell(15, 10).value)
    sf2_t1_sets = [v for v in [ko_ws.cell(15, 12).value, ko_ws.cell(15, 14).value, ko_ws.cell(15, 16).value] if v is not None]
    sf2_t2 = _clean(ko_ws.cell(21, 10).value)
    sf2_t2_sets = [v for v in [ko_ws.cell(21, 12).value, ko_ws.cell(21, 14).value, ko_ws.cell(21, 16).value] if v is not None]
    sf2_winner = _match_winner(sf2_t1, sf2_t1_sets, sf2_t2, sf2_t2_sets)

    # Final: R5 (team1) vs R17 (team2); scores at T/V/X
    final_t1 = _clean(ko_ws.cell(5, 18).value)
    final_t1_sets = [v for v in [ko_ws.cell(5, 20).value, ko_ws.cell(5, 22).value, ko_ws.cell(5, 24).value] if v is not None]
    final_t2 = _clean(ko_ws.cell(17, 18).value)
    final_t2_sets = [v for v in [ko_ws.cell(17, 20).value, ko_ws.cell(17, 22).value, ko_ws.cell(17, 24).value] if v is not None]
    final_winner = _match_winner(final_t1, final_t1_sets, final_t2, final_t2_sets)

    # Winner Z11, Runner-up Z21; no 3rd place score recorded so leave blank
    winner = _clean(ko_ws.cell(11, 26).value)
    runner_up = _clean(ko_ws.cell(21, 26).value)
    third = ""

    # Eliminator winner = whichever team appears in the semi-finals
    sf_teams = {sf1_t1, sf1_t2, sf2_t1, sf2_t2}
    elim_winner = elim_t1 if elim_t1 in sf_teams else elim_t2

    sf1_loser = sf1_t2 if sf1_winner == sf1_t1 else sf1_t1
    sf2_loser = sf2_t2 if sf2_winner == sf2_t1 else sf2_t1

    return {
        "year": 2018,
        "title": "HHB Annual Doubles Classic 2018",
        "rules": rules,
        "important_dates": important_dates,
        "groups": [{"name": "League", "matches": matches, "standings": standings}],
        "knockouts": {
            "qf_label": "Eliminator",
            "quarterfinals": [{
                "team1": elim_t1, "team1_sets": elim_t1_sets,
                "team2": elim_t2, "team2_sets": elim_t2_sets,
                "winner": elim_winner,
            }],
            "semifinals": [
                {"team1": sf1_t1, "team1_sets": sf1_t1_sets, "team2": sf1_t2, "team2_sets": sf1_t2_sets, "winner": sf1_winner},
                {"team1": sf2_t1, "team1_sets": sf2_t1_sets, "team2": sf2_t2, "team2_sets": sf2_t2_sets, "winner": sf2_winner},
            ],
            "final": {
                "team1": final_t1, "team1_sets": final_t1_sets,
                "team2": final_t2, "team2_sets": final_t2_sets,
                "winner": final_winner or winner,
            },
            "third_place": {
                "team1": sf1_loser, "team1_sets": [],
                "team2": sf2_loser, "team2_sets": [],
                "winner": third,
            },
            "winner": winner or final_winner,
            "runner_up": runner_up,
            "third": third,
        },
    }


def get_doubles_tournament(year):
    path = TOURNAMENTS_DIR / _doubles_filename(year)
    if not path.exists():
        return None

    wb = openpyxl.load_workbook(path, data_only=True)

    if year == 2018:
        return _parse_doubles_2018(wb)

    dates_ws = wb["Dates"]

    title = _clean(dates_ws["A1"].value).strip()

    rules = _parse_rules(dates_ws)
    important_dates = _parse_important_dates(dates_ws)

    group_sheet_names = [
        s for s in wb.sheetnames
        if s.lower().startswith("group ") or s in ("Super 6", "Super 3")
    ]
    groups = [_parse_group_sheet(wb[s], s) for s in group_sheet_names]

    knockouts = None
    if "Knockouts" in wb.sheetnames:
        kws = wb["Knockouts"]
        try:
            if "Super 6" in wb.sheetnames:
                # 2022's bespoke qualification + walkover format
                knockouts = _parse_knockouts_2022(kws)
            elif kws["A3"].value == "A1":
                # 4-team direct-to-semis format (2019, 2021)
                knockouts = _parse_knockouts_direct_semis(kws)
            elif kws["A5"].value == "A1":
                # Modern 8-team quarterfinals format (2021 onwards, mostly)
                parsed = _parse_knockouts(kws)
                if parsed["winner"] or parsed["quarterfinals"][0]["team1"]:
                    knockouts = parsed
        except Exception:
            knockouts = None

    return {
        "year": year,
        "title": title,
        "rules": rules,
        "important_dates": important_dates,
        "groups": groups,
        "knockouts": knockouts,
    }

    return {
        "year": year,
        "title": title,
        "rules": rules,
        "important_dates": important_dates,
        "groups": groups,
        "knockouts": knockouts,
    }
